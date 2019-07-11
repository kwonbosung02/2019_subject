import openpyxl
from openpyxl import load_workbook,Workbook


wb = openpyxl.load_workbook('words.xlsx')

ws = wb['Sheet']

write_wb = Workbook()
write_ws = write_wb.active
cnt = 1
write_ws.append(['[문학 단어장 프로그램]'])

if __name__ == "__main__":
    while(True):
        s1 = input("단어 : ")
        s2 = input("뜻 : ")

        write_ws.append([s1,s2])
        write_wb.save('Words.xlsx')



